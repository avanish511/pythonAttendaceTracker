# -*- coding: utf-8 -*-
"""
Created on Fri Jun  2 16:25:27 2023

@author: Dell
"""

import tkinter as tk
from tkinter import messagebox
import openpyxl
import requests
from twilio.rest import Client

# Twilio API configuration
TWILIO_ACCOUNT_SID = 'ACcb3ee86a938f5cbf20f348e505cb160d'
TWILIO_AUTH_TOKEN = '75c437c73ad084056403c408d7aa8085'
TWILIO_PHONE_NUMBER = '+13613102463'

# Sendinblue API configuration
SENDINBLUE_API_KEY = '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'
SENDER_EMAIL = 'stuffynaan@gmail.com'

# Loading the Excel sheet
excel_file = 'D:\\attendace.xlsx'
wb = openpyxl.load_workbook(excel_file)
sheet = wb['Sheet1']

# Getting the max row and column numbers
max_row = sheet.max_row
max_column = sheet.max_column

# Create a Tkinter window
window = tk.Tk()
window.title("Attendance Tracker")

# Function to update attendance
def update_attendance():
    roll_number = int(roll_number_entry.get())
    subject_code = int(subject_combobox.get())
    
    if not roll_number or not subject_code:
        messagebox.showerror("Error", "Roll number and subject code are required!")
        return

    for row in range(2, max_row + 1):
        if sheet.cell(row=row, column=1).value == roll_number:
            column_index = subject_code + 2
            attendance = sheet.cell(row=row, column=column_index).value
            attendance += 1
            sheet.cell(row=row, column=column_index).value = attendance
            break
    else:
        messagebox.showerror("Error", "Roll number not found!")
        return

    wb.save(excel_file)
    messagebox.showinfo("Success", "Attendance updated!")

# Function to send warning email and SMS for lack of attendance
def send_warning():
    subject_code = int(subject_combobox.get())
    subject = get_subject_name(subject_code)
    message = f"Warning! You have lack of attendance in {subject}.You could be debarred from writing SEE"
    roll_numbers = roll_numbers_entry.get()
    emails = emails_entry.get()
    phone_numbers = phone_numbers_entry.get()

    roll_numbers = roll_numbers.split(',')
    emails = emails.split(',')
    phone_numbers = phone_numbers.split(',')

    if len(roll_numbers) != len(emails) or len(roll_numbers) != len(phone_numbers):
        messagebox.showerror("Error", "Number of roll numbers, emails, and phone numbers should be the same!")
        return

    for i in range(len(roll_numbers)):
        roll_number = int(roll_numbers[i])
        email = emails[i].strip()
        phone_number = phone_numbers[i].strip()

        send_email(email, subject, message)
        send_sms(phone_number, message)

# Function to send email using Sendinblue API
def send_email(email, subject, message):
    url = 'https://api.sendinblue.com/v3/smtp/email'
    headers = {
        'Content-Type': 'application/json',
        'api-key': SENDINBLUE_API_KEY
    }
    payload = {
        'sender': {
            'email': SENDER_EMAIL
        },
        'to': [
            {
                'email': email
            }
        ],
        'subject': subject,
        'htmlContent': message
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 201:
            print(f"Email sent to {email}")
        else:
            print(f"Failed to send email to {email}")
    except requests.exceptions.RequestException as e:
        print(f"Error sending email to {email}: {str(e)}")

# Function to send SMS using Twilio API
def send_sms(phone_number, message):
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    try:
        message = client.messages.create(
            body=message,
            from_=TWILIO_PHONE_NUMBER,
            to=phone_number
        )
        print(f"SMS sent to {phone_number}")
    except Exception as e:
        print(f"Error sending SMS to {phone_number}: {str(e)}")

# Function to get subject name based on subject code
def get_subject_name(subject_code):
    subject_names = {1: "Java", 2: "Python", 3: "DSA"}
    return subject_names.get(subject_code, "")

# Create the GUI elements
subject_label = tk.Label(window, text="Subject Code:(1-3)")
subject_label.pack()
subject_combobox = tk.Entry(window)
subject_combobox.pack()

roll_number_label = tk.Label(window, text="Roll Number:")
roll_number_label.pack()
roll_number_entry = tk.Entry(window)
roll_number_entry.pack()

update_button = tk.Button(window, text="Update Attendance", command=update_attendance)
update_button.pack()

warning_label = tk.Label(window, text="Send Warning:")
warning_label.pack()
roll_numbers_label = tk.Label(window, text="Roll Numbers (comma-separated):")
roll_numbers_label.pack()
roll_numbers_entry = tk.Entry(window)
roll_numbers_entry.pack()

emails_label = tk.Label(window, text="Emails (comma-separated):")
emails_label.pack()
emails_entry = tk.Entry(window)
emails_entry.pack()

phone_numbers_label = tk.Label(window, text="Phone Numbers (comma-separated) with country code:")
phone_numbers_label.pack()
phone_numbers_entry = tk.Entry(window)
phone_numbers_entry.pack()

send_warning_button = tk.Button(window, text="Send Warning", command=send_warning)
send_warning_button.pack()

window.mainloop()
